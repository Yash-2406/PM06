"""ExportService — Excel export and MIS summary data.

Writes the tracker Excel in `New_Connection_FY26.xlsx` format
and provides MIS aggregation helpers.
"""

from __future__ import annotations

import logging
import shutil
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.data.case_repository import CaseRepository
from app.data.database import Database
from app.domain.constants import TRACKER_COLUMNS
from app.domain.enums import CaseStatus
from app.domain.models import Case, TrackerRow
from app.infrastructure.config_manager import ConfigManager
from app.infrastructure.formatting import format_indian_amount

logger = logging.getLogger(__name__)


# ── Shared helper ───────────────────────────────────────────────


def case_to_tracker_row(case: Case, sl_no: int) -> TrackerRow:
    """Convert a Case to a TrackerRow.

    Shared by ExportService and TrackerService so the mapping is
    defined in exactly one place.
    """
    now_str = datetime.now().strftime("%d-%m-%Y")
    created_str = (
        case.created_at.strftime("%d-%m-%Y") if case.created_at else now_str
    )
    processed_str = (
        case.updated_at.strftime("%d-%m-%Y") if case.updated_at else now_str
    )
    amount_str = (
        format_indian_amount(case.grand_total) if case.grand_total else "0"
    )
    return TrackerRow(
        sl_no=sl_no,
        scheme_no=case.order_no or "",
        n_no=case.notification_no or "",
        district=case.district or "",
        zone=case.zone or "",
        date_received=created_str,
        date_processed=processed_str,
        status=case.status.value if case.status else "PENDING",
        remarks="",
        amount_rs=amount_str,
        correction_suggested="Yes" if case.correction_details else "No",
        correction_details=case.correction_details or "",
    )


class ExportService:
    """Handles data export to Excel and MIS reporting."""

    def __init__(
        self,
        db: Database | None = None,
        config: ConfigManager | None = None,
    ) -> None:
        self._config = config or ConfigManager()
        self._db = db or Database(self._config.db_path)
        self._repo = CaseRepository(self._db)

    # ── Excel Export ────────────────────────────────────────────

    def export_to_excel(
        self,
        output_path: str | Path,
        district: Optional[str] = None,
        zone: Optional[str] = None,
        status: Optional[CaseStatus] = None,
        date_from: Optional[str] = None,
        date_to: Optional[str] = None,
        progress_cb: Optional[Callable[[int, int], None]] = None,
    ) -> Path:
        """Export filtered cases to an Excel file matching tracker format.

        Args:
            output_path: Target .xlsx file path.
            district: Filter by district code (None = all).
            zone: Filter by zone code (None = all).
            status: Filter by CaseStatus enum (None = all).
            date_from: Filter by date received >= (yyyy-mm-dd, None = no limit).
            date_to: Filter by date received <= (yyyy-mm-dd, None = no limit).
            progress_cb: Optional callback(current_row, total_rows) for progress.
        """
        output_path = Path(output_path)

        # Convert CaseStatus enum → string value for the repository query
        status_str = status.value if status else None

        cases = self._repo.list_all(
            district=district,
            zone=zone,
            status=status_str,
            date_from=date_from,
            date_to=date_to,
        )
        total = len(cases)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tracker"

        # Header row
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=11, color="FFFFFF")

        for col_idx, col_name in enumerate(TRACKER_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Data rows — sequential serial numbers starting from 1
        for sl_no, case in enumerate(cases, start=1):
            row = case_to_tracker_row(case, sl_no=sl_no)
            row_data = [
                row.sl_no,
                row.scheme_no,
                row.n_no,
                row.district,
                row.zone,
                row.date_received,
                row.date_processed,
                row.status,
                row.remarks,
                row.amount_rs,
                row.correction_suggested,
                row.correction_details,
            ]
            excel_row = sl_no + 1  # +1 for header
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=excel_row, column=col_idx, value=value)

            if progress_cb and sl_no % 50 == 0:
                progress_cb(sl_no, total)

        # Auto-fit column widths (approximate)
        for col_idx in range(1, len(TRACKER_COLUMNS) + 1):
            col_letter = get_column_letter(col_idx)
            max_len = len(str(TRACKER_COLUMNS[col_idx - 1]))
            for row in ws.iter_rows(
                min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx
            ):
                for cell in row:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

        # Atomic write: save to temp file then move to final path
        output_path.parent.mkdir(parents=True, exist_ok=True)
        tmp_fd, tmp_path = tempfile.mkstemp(
            dir=str(output_path.parent), suffix=".xlsx"
        )
        try:
            import os
            os.close(tmp_fd)
            wb.save(tmp_path)
            wb.close()
            shutil.move(tmp_path, str(output_path))
        except BaseException:
            try:
                Path(tmp_path).unlink(missing_ok=True)
            except OSError:
                pass
            wb.close()
            raise

        if progress_cb:
            progress_cb(total, total)

        logger.info("Exported %d cases to %s", total, output_path)
        return output_path

    # ── MIS Summary ─────────────────────────────────────────────

    def get_mis_data(self) -> Dict[str, Any]:
        """Return MIS aggregation data for the MIS tab.

        Uses SQL aggregation queries instead of loading all cases
        into memory.
        """
        district_status_counts = self._repo.count_by_district_status()

        total_cases, total_amount = self._repo.count_and_sum_all()
        by_status = self._repo.count_by_status()
        by_district = self._repo.count_by_district()
        by_zone = self._repo.count_by_zone()
        by_work_type = self._repo.count_by_work_type()
        amount_by_district = self._repo.sum_by_district()
        amount_by_status = self._repo.sum_by_status()
        monthly_trend = self._repo.count_by_month()

        return {
            "district_status_counts": district_status_counts,
            "total_cases": total_cases,
            "total_amount": total_amount,
            "total_amount_formatted": format_indian_amount(total_amount),
            "by_status": by_status,
            "by_district": by_district,
            "by_zone": by_zone,
            "by_work_type": by_work_type,
            "amount_by_district": amount_by_district,
            "amount_by_status": amount_by_status,
            "monthly_trend": monthly_trend,
        }

    def export_mis_to_excel(self, output_path: str | Path) -> Path:
        """Export MIS summary data to an Excel workbook."""
        output_path = Path(output_path)
        data = self.get_mis_data()

        wb = openpyxl.Workbook()

        # ── Sheet 1: Overview ───────────────────────────────────
        ws = wb.active
        ws.title = "Overview"
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_w = Font(bold=True, size=11, color="FFFFFF")

        ws.append(["MIS Summary Report"])
        ws["A1"].font = Font(bold=True, size=14)
        ws.append([f"Generated: {datetime.now().strftime('%d-%b-%Y %H:%M')}"])
        ws.append([])
        ws.append(["Total Cases", data["total_cases"]])
        ws.append(["Total Amount (₹)", data["total_amount"]])
        ws["B5"].number_format = '#,##0.00'

        # ── Sheet 2: By Status ──────────────────────────────────
        ws2 = wb.create_sheet("By Status")
        ws2.append(["Status", "Count", "Amount (₹)"])
        for cell in ws2[1]:
            cell.font = header_font_w
            cell.fill = header_fill
        amount_by_status = data.get("amount_by_status", {})
        for status, count in sorted(data["by_status"].items()):
            ws2.append([status, count, amount_by_status.get(status, 0)])

        # ── Sheet 3: By District ────────────────────────────────
        ws3 = wb.create_sheet("By District")
        ws3.append(["District", "Count", "Amount (₹)"])
        for cell in ws3[1]:
            cell.font = header_font_w
            cell.fill = header_fill
        amount_by_district = data.get("amount_by_district", {})
        for district, count in sorted(data["by_district"].items()):
            ws3.append([district, count, amount_by_district.get(district, 0)])

        # ── Sheet 4: By Zone ────────────────────────────────────
        ws4 = wb.create_sheet("By Zone")
        ws4.append(["Zone", "Count"])
        for cell in ws4[1]:
            cell.font = header_font_w
            cell.fill = header_fill
        for zone, count in sorted(data["by_zone"].items()):
            ws4.append([zone, count])

        # ── Sheet 5: By Work Type ───────────────────────────────
        ws5 = wb.create_sheet("By Work Type")
        ws5.append(["Work Type", "Count"])
        for cell in ws5[1]:
            cell.font = header_font_w
            cell.fill = header_fill
        for wt, count in sorted(data["by_work_type"].items()):
            ws5.append([wt, count])

        # ── Sheet 6: Monthly Trend ──────────────────────────────
        ws6 = wb.create_sheet("Monthly Trend")
        ws6.append(["Month", "Count", "Amount (₹)"])
        for cell in ws6[1]:
            cell.font = header_font_w
            cell.fill = header_fill
        for row in data.get("monthly_trend", []):
            ws6.append([row["month"], row["cnt"], row["total"]])

        # ── Sheet 7: District × Status Matrix ───────────────────
        ws7 = wb.create_sheet("District × Status")
        statuses = sorted({r["status"] for r in data["district_status_counts"]})
        ws7.append(["District"] + statuses + ["Total"])
        for cell in ws7[1]:
            cell.font = header_font_w
            cell.fill = header_fill
        # Build matrix
        matrix: Dict[str, Dict[str, int]] = {}
        for r in data["district_status_counts"]:
            d = r.get("district_code") or "UNKNOWN"
            s = r.get("status") or "UNKNOWN"
            matrix.setdefault(d, {})[s] = r["count"]
        for district in sorted(matrix.keys()):
            row_vals = [district]
            row_total = 0
            for s in statuses:
                c = matrix[district].get(s, 0)
                row_vals.append(c)
                row_total += c
            row_vals.append(row_total)
            ws7.append(row_vals)

        # Auto-fit columns for all sheets
        for sheet in wb.worksheets:
            for col in sheet.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        max_len = max(max_len, len(str(cell.value or "")))
                    except Exception:
                        pass
                sheet.column_dimensions[col_letter].width = min(max_len + 4, 40)

        # Atomic write
        tmp_dir = output_path.parent
        with tempfile.NamedTemporaryFile(
            dir=tmp_dir, suffix=".xlsx", delete=False
        ) as tmp:
            tmp_path = Path(tmp.name)
        try:
            wb.save(str(tmp_path))
            wb.close()
            shutil.move(str(tmp_path), str(output_path))
        except Exception:
            tmp_path.unlink(missing_ok=True)
            wb.close()
            raise

        logger.info("Exported MIS summary to %s", output_path)
        return output_path