"""Excel repository — read/write the New_Connection_FY26.xlsx tracker.

Atomic writes with PermissionError handling for when Excel has the file open.
"""

from __future__ import annotations

import shutil
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Optional

from filelock import FileLock

from app.domain.constants import TRACKER_COLUMNS
from app.domain.exceptions import TrackerWriteError
from app.domain.models import TrackerRow
from app.infrastructure.logger import get_logger

logger = get_logger(__name__)


class ExcelRepository:
    """Read/write the New_Connection_FY26.xlsx tracker file."""

    def __init__(self, tracker_path: Path) -> None:
        self._tracker_path = tracker_path
        self._lock_path = tracker_path.with_suffix(".lock")

    def _ensure_tracker_exists(self) -> None:
        """Create the tracker file with headers if it doesn't exist."""
        if self._tracker_path.exists():
            return
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        for col_idx, col_name in enumerate(TRACKER_COLUMNS, start=1):
            ws.cell(row=1, column=col_idx, value=col_name)
        self._tracker_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(self._tracker_path))
        wb.close()
        logger.info("Created new tracker file: %s", self._tracker_path)

    def get_max_sl_no(self) -> int:
        """Return the current maximum Sl. No. in the tracker."""
        self._ensure_tracker_exists()
        import openpyxl

        wb = openpyxl.load_workbook(str(self._tracker_path), read_only=True)
        ws = wb["Sheet1"]
        max_sl = 0
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0] is not None:
                try:
                    max_sl = max(max_sl, int(row[0]))
                except (ValueError, TypeError):
                    pass
        wb.close()
        return max_sl

    def find_by_scheme_no(self, scheme_no: str) -> Optional[int]:
        """Find row number of existing entry with given Scheme no.

        Returns the Excel row number (1-based) or None.
        """
        self._ensure_tracker_exists()
        import openpyxl

        wb = openpyxl.load_workbook(str(self._tracker_path), read_only=True)
        ws = wb["Sheet1"]
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[1] is not None and str(row[1]).strip() == scheme_no:
                wb.close()
                return row_idx
        wb.close()
        return None

    def append_row(self, tracker_row: TrackerRow) -> None:
        """Append a new row to the tracker file.

        Uses file locking and atomic write pattern.

        Raises:
            TrackerWriteError: If the file is locked by Excel or write fails.
        """
        self._ensure_tracker_exists()
        lock = FileLock(str(self._lock_path), timeout=5)
        try:
            with lock:
                self._write_row(tracker_row, update_row=None)
        except PermissionError as e:
            raise TrackerWriteError(
                f"Cannot write to tracker: {e}",
                user_message="The tracker Excel file is open in another program. "
                "Please close it and click 'Sync' to retry.",
            ) from e
        except TimeoutError as e:
            raise TrackerWriteError(
                f"Tracker file lock timeout: {e}",
                user_message="The tracker file is being used by another process. "
                "Please wait and try again.",
            ) from e

    def update_row(self, row_number: int, tracker_row: TrackerRow) -> None:
        """Update an existing row in the tracker file.

        Args:
            row_number: 1-based Excel row number to update.
            tracker_row: New data for the row.

        Raises:
            TrackerWriteError: If write fails.
        """
        self._ensure_tracker_exists()
        lock = FileLock(str(self._lock_path), timeout=5)
        try:
            with lock:
                self._write_row(tracker_row, update_row=row_number)
        except PermissionError as e:
            raise TrackerWriteError(
                f"Cannot write to tracker: {e}",
                user_message="The tracker Excel file is open in another program. "
                "Please close it and click 'Sync' to retry.",
            ) from e

    def append_or_update_row(self, tracker_row: TrackerRow) -> None:
        """Append a row, or update an existing row if a matching scheme_no exists."""
        existing = self._find_row_by_scheme(tracker_row.scheme_no)
        if existing is not None:
            self.update_row(existing, tracker_row)
        else:
            self.append_row(tracker_row)

    def _find_row_by_scheme(self, scheme_no: str) -> Optional[int]:
        """Return the 1-based row number of the first row matching scheme_no, or None."""
        if not scheme_no or not self._tracker_path.exists():
            return None
        import openpyxl

        wb = openpyxl.load_workbook(str(self._tracker_path), read_only=True)
        ws = wb["Sheet1"]
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            # scheme_no is in column B (index 1)
            if len(row) > 1 and row[1].value and str(row[1].value).strip() == scheme_no.strip():
                wb.close()
                return row_idx
        wb.close()
        return None

    def _write_row(self, tracker_row: TrackerRow, update_row: Optional[int]) -> None:
        """Internal write implementation with atomic save."""
        import openpyxl

        wb = openpyxl.load_workbook(str(self._tracker_path))
        ws = wb["Sheet1"]

        row_data = [
            tracker_row.sl_no,
            tracker_row.scheme_no,
            tracker_row.n_no,
            tracker_row.district,
            tracker_row.zone,
            tracker_row.date_received,
            tracker_row.date_processed,
            tracker_row.status,
            tracker_row.remarks,
            tracker_row.amount_rs,
            tracker_row.correction_suggested,
            tracker_row.correction_details,
        ]

        if update_row is not None:
            target_row = update_row
        else:
            target_row = ws.max_row + 1

        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=target_row, column=col_idx, value=value)

        # Atomic save: write to temp, then replace
        tmp_fd, tmp_path = tempfile.mkstemp(
            dir=str(self._tracker_path.parent),
            suffix=".xlsx",
        )
        try:
            import os
            os.close(tmp_fd)
            wb.save(tmp_path)
            wb.close()
            shutil.move(tmp_path, str(self._tracker_path))
            logger.info("Tracker row written: Scheme=%s", tracker_row.scheme_no)
        except BaseException:
            try:
                Path(tmp_path).unlink(missing_ok=True)
            except OSError:
                pass
            wb.close()

    # ── Batch write (single open/save cycle) ────────────────────

    def batch_write_rows(self, rows: list[TrackerRow]) -> None:
        """Write many rows in a single open → write → save cycle.

        For each row, if a matching scheme_no already exists it is
        updated in-place; otherwise the row is appended.  The file is
        opened and saved only once, avoiding the O(N) open/save
        overhead of calling append_or_update_row() per row.

        Raises:
            TrackerWriteError: If the file is locked or write fails.
        """
        if not rows:
            return
        self._ensure_tracker_exists()
        lock = FileLock(str(self._lock_path), timeout=5)
        try:
            with lock:
                self._batch_write_impl(rows)
        except PermissionError as e:
            raise TrackerWriteError(
                f"Cannot write to tracker: {e}",
                user_message="The tracker Excel file is open in another program. "
                "Please close it and click 'Sync' to retry.",
            ) from e
        except TimeoutError as e:
            raise TrackerWriteError(
                f"Tracker file lock timeout: {e}",
                user_message="The tracker file is being used by another process. "
                "Please wait and try again.",
            ) from e

    def _batch_write_impl(self, rows: list[TrackerRow]) -> None:
        """Internal: open workbook once, upsert all rows, save once."""
        import openpyxl

        wb = openpyxl.load_workbook(str(self._tracker_path))
        ws = wb["Sheet1"]

        # Build scheme_no → row_number index for existing data
        existing: dict[str, int] = {}
        for row_idx in range(2, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=2).value
            if val is not None:
                existing[str(val).strip()] = row_idx

        next_row = ws.max_row + 1

        for tracker_row in rows:
            row_data = [
                tracker_row.sl_no,
                tracker_row.scheme_no,
                tracker_row.n_no,
                tracker_row.district,
                tracker_row.zone,
                tracker_row.date_received,
                tracker_row.date_processed,
                tracker_row.status,
                tracker_row.remarks,
                tracker_row.amount_rs,
                tracker_row.correction_suggested,
                tracker_row.correction_details,
            ]
            target = existing.get(tracker_row.scheme_no.strip())
            if target is None:
                target = next_row
                next_row += 1
                existing[tracker_row.scheme_no.strip()] = target

            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=target, column=col_idx, value=value)

        # Atomic save
        tmp_fd, tmp_path = tempfile.mkstemp(
            dir=str(self._tracker_path.parent), suffix=".xlsx"
        )
        try:
            import os
            os.close(tmp_fd)
            wb.save(tmp_path)
            wb.close()
            shutil.move(tmp_path, str(self._tracker_path))
            logger.info("Batch wrote %d tracker rows", len(rows))
        except BaseException:
            try:
                Path(tmp_path).unlink(missing_ok=True)
            except OSError:
                pass
            wb.close()
            raise
            raise

    def read_all_rows(self) -> list[dict]:
        """Read all data rows from the tracker file."""
        self._ensure_tracker_exists()
        import openpyxl

        wb = openpyxl.load_workbook(str(self._tracker_path), read_only=True)
        ws = wb["Sheet1"]
        rows: list[dict] = []
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                row_dict = {}
                for idx, header in enumerate(headers):
                    row_dict[header] = row[idx] if idx < len(row) else None
                rows.append(row_dict)
        wb.close()
        return rows

    def save_backup(self, backup_dir: Path) -> Path:
        """Save a timestamped backup of the tracker file."""
        if not self._tracker_path.exists():
            raise FileNotFoundError(f"Tracker file not found: {self._tracker_path}")
        backup_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = backup_dir / f"tracker_backup_{timestamp}.xlsx"
        shutil.copy2(str(self._tracker_path), str(backup_path))
        logger.info("Tracker backup saved: %s", backup_path)
        return backup_path
