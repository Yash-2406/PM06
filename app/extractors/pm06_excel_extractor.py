"""PM06 Format Excel extractor — rules PM-1 through PM-7.

MANDATORY document. Most reliable data source.
NEVER uses hard-coded row numbers.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any, Optional

import openpyxl

from app.domain.constants import PM06_SHEET_NAME, SCOPE_KEYWORDS
from app.domain.enums import FieldConfidence, FileType
from app.domain.models import ExtractionResult, FeederDetail
from app.extractors.base_extractor import BaseExtractor
from app.infrastructure.file_utils import validate_file_type
from app.infrastructure.logger import get_logger
from app.infrastructure.text_utils import normalise_dt_capacity, normalise_label

logger = get_logger(__name__)


class PM06ExcelExtractor(BaseExtractor):
    """Extract structured intake data from the PM06 Format Excel file.

    Builds a label-value map by scanning ALL rows (PM-3).
    Never uses hard-coded row numbers (PM-2).
    """

    def _validate_file(self, file_path: Path) -> None:
        super()._validate_file(file_path)
        validate_file_type(file_path, FileType.PM06_EXCEL)

    def _do_extract(self, file_path: Path) -> dict[str, ExtractionResult]:
        results: dict[str, ExtractionResult] = {}

        wb = openpyxl.load_workbook(str(file_path), data_only=True)

        # PM-1: Find sheet by stripped name
        ws = self._find_format_sheet(wb)
        if ws is None:
            wb.close()
            sheet_names = ", ".join(wb.sheetnames)
            return self._error_result(
                f"Cannot find 'Format' sheet. Found: {sheet_names}. "
                f"Please upload the correct PM06 Format Excel file."
            )

        # PM-3: Build label-value map
        label_map = self._build_label_value_map(ws)

        # Extract all fields via PM-4 fuzzy keyword matching
        results["order_no"] = self._extract_field(label_map, ["order"], "Order No")
        notif_result = self._extract_field(label_map, ["request"], "Request No")
        # Strip common prefixes like "NN. " from notification number
        if notif_result.value and isinstance(notif_result.value, str):
            notif_result.value = re.sub(r"^(?:NN\.?\s*)", "", notif_result.value).strip()
        results["notification_no"] = notif_result
        results["applicant_name"] = self._extract_field(label_map, ["consumer", "name"], "Consumer Name")
        results["address"] = self._extract_field(label_map, ["address"], "Address")
        results["sanctioned_load"] = self._extract_field(label_map, ["load"], "Sanctioned Load")
        results["area_type"] = self._extract_field(label_map, ["area", "type"], "Area Type")
        results["proposal_type"] = self._extract_field(label_map, ["proposal"], "Type of Proposal")
        results["dt_code"] = self._extract_field(label_map, ["dt", "code"], "DT Code")
        results["substation_name"] = self._extract_field(label_map, ["station", "name"], "Sub Station Name")
        results["tapping_pole"] = self._extract_tapping_pole(ws, label_map)
        results["detailed_reason"] = self._extract_field(label_map, ["reason"], "Detailed Reason")

        # Existing DT Capacity — with normalisation (PM-7)
        dt_raw = self._find_field(label_map, ["existing", "dt"])
        if dt_raw is None:
            # Fallback: some sheets label it "DT Capacity" without "Existing"
            dt_raw = self._find_field(label_map, ["dt", "capacity"])
        if dt_raw is not None:
            normalised = normalise_dt_capacity(str(dt_raw))
            results["dt_capacity_existing"] = self._make_result(
                normalised, source="label_map"
            )
        else:
            results["dt_capacity_existing"] = self._not_found("Existing DT Capacity")

        # DT Loading
        results["dt_loading"] = self._extract_field(label_map, ["dt", "loading"], "DT Loading")

        # Number of feeders
        results["num_feeders"] = self._extract_field(label_map, ["number", "feeder"], "Number of LT Feeders")

        # PM-5: Scope of Work — three-strategy extraction
        scope = self._find_scope_of_work(ws, label_map)
        if scope:
            results["scope_of_work"] = self._make_result(scope, source="label_map")
        else:
            results["scope_of_work"] = self._not_found(
                "Scope of Work", "Not found — will auto-generate from BOM"
            )

        # PM-6: Feeder details
        feeders = self._extract_feeder_details(ws)
        results["feeders"] = self._make_result(feeders, source="feeder_table")

        # PM-8: LT line extension materials (Material, Quantity sub-table)
        lt_ext_mats = self._extract_lt_extension_materials(ws)
        if lt_ext_mats:
            results["lt_extension_materials"] = self._make_result(
                lt_ext_mats, source="lt_extension_table"
            )

        wb.close()
        return results

    # ------------------------------------------------------------------
    # PM-1: Find the "Format " sheet (handle trailing space)
    # ------------------------------------------------------------------

    @staticmethod
    def _find_format_sheet(wb: openpyxl.Workbook) -> Any:
        """Find the Format sheet by comparing stripped names.

        Fallback: if no sheet named 'Format', use the first sheet whose
        row-1 cell A1 contains 'format of lt' (common variant names like
        'Sheet1', '1', etc.).
        """
        for name in wb.sheetnames:
            if name.strip().lower() == PM06_SHEET_NAME.lower():
                return wb[name]
        # Fallback: check row 1 of each sheet for the PM06 header
        for name in wb.sheetnames:
            ws = wb[name]
            a1 = ws.cell(1, 1).value
            if a1 and "format of lt" in str(a1).lower():
                return ws
        return None

    # ------------------------------------------------------------------
    # PM-3: Build label-value map by scanning ALL rows
    # ------------------------------------------------------------------

    @staticmethod
    def _build_label_value_map(ws: Any) -> dict[str, Any]:
        """Scan every row. Labels in column A or B, value is next non-empty cell."""
        label_map: dict[str, Any] = {}
        for row in ws.iter_rows(values_only=True):
            for col_idx in (0, 1):
                cell_val = row[col_idx] if len(row) > col_idx else None
                if cell_val and isinstance(cell_val, str) and len(str(cell_val).strip()) > 1:
                    normalised = normalise_label(str(cell_val))
                    value = None
                    for val_idx in range(col_idx + 1, min(len(row), col_idx + 5)):
                        if row[val_idx] is not None:
                            value = row[val_idx]
                            break
                    if normalised and value is not None:
                        if normalised not in label_map:
                            label_map[normalised] = value
        return label_map

    # ------------------------------------------------------------------
    # PM-4: Fuzzy keyword matching
    # ------------------------------------------------------------------

    @staticmethod
    def _find_field(label_map: dict[str, Any], keywords: list[str]) -> Optional[Any]:
        """Return value if ALL keywords appear in the normalised label."""
        for label, value in label_map.items():
            if all(kw in label for kw in keywords):
                return value
        return None

    def _extract_field(
        self, label_map: dict[str, Any], keywords: list[str], field_name: str
    ) -> ExtractionResult:
        """Extract a field using fuzzy matching, wrapping in ExtractionResult."""
        value = self._find_field(label_map, keywords)
        if value is not None:
            return self._make_result(str(value).strip(), source="label_map")
        return self._not_found(field_name)

    def _extract_tapping_pole(
        self, ws: Any, label_map: dict[str, Any]
    ) -> ExtractionResult:
        """Extract tapping pole — may be in a later column than the generic value.

        Row format: [sr, 'Tapping Point', 'Pole', 'HT572-63/21A', None]
        The label map picks 'Pole' (generic), but we need 'HT572-63/21A'.
        """
        # Pattern: pole codes like HT572-63/21A, UHT572-27, U511-49/17, 511-65/5, 523-53/1/1
        _POLE_CODE = re.compile(r"(?:U?HT)?\d{3}[-/]", re.IGNORECASE)
        # Generic words that are NOT valid pole codes (checked per-word)
        _GARBAGE_WORDS = {"pole", "number", "dt", "none", "na", "tapping", "point", "name"}

        def _is_garbage(val: str) -> bool:
            """True if value is generic garbage, not a real pole code."""
            words = val.lower().split()
            return all(w in _GARBAGE_WORDS for w in words)

        # Strategy 1: scan rows for "tapping" label and look for pole-code value
        for row in ws.iter_rows(values_only=True):
            has_tapping = False
            for cell in row:
                if cell and isinstance(cell, str) and "tapping" in cell.lower():
                    has_tapping = True
                    break
            if has_tapping:
                # Look for HT/numeric pole code in any cell of this row
                for cell in row:
                    if cell and isinstance(cell, str) and _POLE_CODE.search(cell):
                        cleaned = cell.strip().rstrip(".")
                        return self._make_result(cleaned, source="label_map")
                # Fallback: last non-None string cell that is a valid value
                for cell in reversed(list(row)):
                    if cell and isinstance(cell, str):
                        val = cell.strip().rstrip(".")
                        if len(val) > 3 and not _is_garbage(val):
                            return self._make_result(val, source="label_map")

        # Strategy 2: fall back to label map
        value = self._find_field(label_map, ["tapping"])
        if value is not None:
            val_str = str(value).strip().rstrip(".")
            if not _is_garbage(val_str):
                return self._make_result(val_str, source="label_map")

        # Strategy 3: mine pole code from substation, scope, or reason text
        for field_keys in (["station", "name"], ["scope"], ["reason"]):
            text = self._find_field(label_map, field_keys)
            if text and isinstance(text, str):
                m = _POLE_CODE.search(text)
                if m:
                    # Extract the full pole code token at this position
                    start = m.start()
                    token = re.search(r"[A-Za-z0-9/\-]+", text[start:])
                    if token:
                        pole = token.group(0).rstrip(".")
                        if len(pole) > 3 and not _is_garbage(pole):
                            return self._make_result(pole, source="label_map")

        return self._not_found("Tapping Point")

    # ------------------------------------------------------------------
    # PM-5: Scope of Work — three-strategy extraction
    # ------------------------------------------------------------------

    @staticmethod
    def _find_scope_of_work(ws: Any, label_map: dict[str, Any]) -> Optional[str]:
        """Three strategies to extract Scope of Work."""
        # Strategy 1: label-map lookup
        for label, value in label_map.items():
            if "scope" in label:
                val_str = str(value).strip()
                if len(val_str) > 5:
                    return val_str

        # Strategy 2: scan cells for "scope" keyword proximity
        for row in ws.iter_rows(values_only=True):
            for i, cell in enumerate(row):
                if cell and isinstance(cell, str) and "scope" in cell.lower():
                    for j in range(i + 1, min(len(row), i + 5)):
                        if row[j] and isinstance(row[j], str) and len(str(row[j])) > 10:
                            return str(row[j]).strip()

        # Strategy 3: longest sentence with engineering keywords
        all_candidates: list[str] = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell and isinstance(cell, str) and len(cell) > 20:
                    if any(kw in cell.lower() for kw in SCOPE_KEYWORDS):
                        all_candidates.append(cell)
        if all_candidates:
            return max(all_candidates, key=len).strip()

        return None

    # ------------------------------------------------------------------
    # PM-6: Feeder sub-table — keyword-bounded scan
    # ------------------------------------------------------------------

    @staticmethod
    def _extract_feeder_details(ws: Any) -> list[FeederDetail]:
        """Extract feeder details from the sub-table bounded by keywords."""
        feeders: list[FeederDetail] = []
        in_section = False
        for row in ws.iter_rows(values_only=True):
            row_text = " ".join(str(c) for c in row if c is not None).lower()
            if "acb" in row_text and ("loading" in row_text or "amps" in row_text):
                in_section = True
                continue
            if in_section:
                if any(kw in row_text for kw in ["tapping", "length", "scope", "reason"]):
                    break
                nums = [c for c in row if isinstance(c, (int, float))]
                if len(nums) >= 2:
                    feeders.append(FeederDetail(
                        sr_no=nums[0],
                        acb_no=nums[1],
                        loading_amps=nums[2] if len(nums) > 2 else None,
                    ))
        return feeders

    # ------------------------------------------------------------------
    # PM-8: LT line extension materials sub-table
    # ------------------------------------------------------------------

    @staticmethod
    def _extract_lt_extension_materials(ws: Any) -> list[dict[str, Any]]:
        """Extract materials from the 'Length of LT line extension' sub-table.

        Rows have: Sr.No. | Material | Quantity (quantity may be blank).
        Returns list of dicts with 'description' and optional 'quantity'.
        """
        materials: list[dict[str, Any]] = []
        in_section = False
        for row in ws.iter_rows(values_only=True):
            row_text = " ".join(str(c) for c in row if c is not None).lower()
            if "length" in row_text and "lt" in row_text and ("extension" in row_text or "line" in row_text):
                in_section = True
                continue
            if in_section:
                # End of section: next labelled row
                if any(kw in row_text for kw in ["reason", "scope", "tapping"]):
                    break
                # Look for material description in columns C/D (index 2/3)
                desc = None
                qty = None
                for cell in row:
                    if cell and isinstance(cell, str) and len(cell.strip()) > 2:
                        # Skip header-like cells and serial numbers
                        if cell.strip().lower() not in ("sr.no.", "material", "quantity"):
                            desc = cell.strip()
                    elif cell and isinstance(cell, (int, float)) and not desc:
                        continue  # serial number before description
                    elif cell and isinstance(cell, (int, float)) and desc:
                        qty = float(cell)
                if desc:
                    materials.append({"description": desc, "quantity": qty})
        return materials