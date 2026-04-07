"""Service-layer tests — TrackerService, ExportService, GeneratorService."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest

from app.data.case_repository import CaseRepository
from app.data.database import Database
from app.domain.enums import CaseStatus, WorkType
from app.domain.models import Case, Material, TrackerRow
from app.services.export_service import ExportService, case_to_tracker_row
from app.services.tracker_service import TrackerService


# ── Fixtures ────────────────────────────────────────────────────


@pytest.fixture
def tracker_svc(db, tmp_path):
    from app.infrastructure.config_manager import ConfigManager
    ConfigManager._instance = None
    config = ConfigManager()
    config._config.set("General", "tracker_path", str(tmp_path / "tracker.xlsx"))
    svc = TrackerService(db=db, config=config)
    yield svc
    ConfigManager._instance = None


@pytest.fixture
def export_svc(db, tmp_path):
    from app.infrastructure.config_manager import ConfigManager
    ConfigManager._instance = None
    config = ConfigManager()
    config._config.set("General", "output_dir", str(tmp_path / "output"))
    svc = ExportService(db=db, config=config)
    yield svc
    ConfigManager._instance = None


def _make_case(**overrides) -> Case:
    defaults = dict(
        order_no="60038419",
        notification_no="1234567890",
        applicant_name="Test User",
        address="123 Test Street, Delhi",
        pin_code="110001",
        zone="411",
        district="CVL",
        wbs_no="CE/N0000/00134",
        load_applied="5 kW",
        category="DOMESTIC",
        work_type=WorkType.LT_STANDARD,
        grand_total=125000.50,
        scope_of_work="LT extension from pole 123 towards premises",
        materials=[Material(description="Cable", quantity=100, unit="MTR")],
        status=CaseStatus.PENDING,
    )
    defaults.update(overrides)
    return Case(**defaults)


# ── case_to_tracker_row ─────────────────────────────────────────


class TestCaseToTrackerRow:
    def test_basic_conversion(self):
        case = _make_case()
        row = case_to_tracker_row(case, sl_no=1)
        assert isinstance(row, TrackerRow)
        assert row.sl_no == 1
        assert row.scheme_no == "60038419"
        assert row.n_no == "1234567890"
        assert row.district == "CVL"
        assert row.zone == "411"
        assert row.status == "Pending"

    def test_amount_formatted(self):
        case = _make_case(grand_total=1234567.89)
        row = case_to_tracker_row(case, sl_no=1)
        assert "1,234,567" in row.amount_rs

    def test_none_grand_total(self):
        case = _make_case(grand_total=None)
        row = case_to_tracker_row(case, sl_no=1)
        assert row.amount_rs == "0"

    def test_correction_suggested_yes(self):
        case = _make_case(correction_details="Wrong zone")
        row = case_to_tracker_row(case, sl_no=1)
        assert row.correction_suggested == "Yes"
        assert row.correction_details == "Wrong zone"

    def test_correction_suggested_no(self):
        case = _make_case(correction_details=None)
        row = case_to_tracker_row(case, sl_no=1)
        assert row.correction_suggested == "No"


# ── TrackerService ──────────────────────────────────────────────


class TestTrackerServiceCRUD:
    def test_list_empty(self, tracker_svc):
        assert tracker_svc.list_cases() == []

    def test_get_nonexistent(self, tracker_svc):
        assert tracker_svc.get_case(9999) is None

    def test_get_by_order_no_nonexistent(self, tracker_svc):
        assert tracker_svc.get_case_by_order_no("99999999") is None

    def test_get_mis_summary(self, tracker_svc):
        summary = tracker_svc.get_mis_summary()
        assert isinstance(summary, (dict, list))


class TestTrackerServiceLifecycle:
    def _create_case(self, db) -> int:
        repo = CaseRepository(db)
        case = _make_case()
        return repo.create_case(case)

    def test_approve_case(self, tracker_svc, db):
        case_id = self._create_case(db)
        tracker_svc.approve_case(case_id, remarks="Looks good")
        case = tracker_svc.get_case(case_id)
        assert case.status == CaseStatus.APPROVED

    def test_reject_case(self, tracker_svc, db):
        case_id = self._create_case(db)
        tracker_svc.reject_case(case_id, "Wrong zone mapping", remarks="Need fix")
        case = tracker_svc.get_case(case_id)
        assert case.status == CaseStatus.REJECTED

    def test_list_cases_with_filter(self, tracker_svc, db):
        self._create_case(db)
        cases = tracker_svc.list_cases(district="CVL")
        assert len(cases) >= 1

    def test_list_cases_wrong_district(self, tracker_svc, db):
        self._create_case(db)
        cases = tracker_svc.list_cases(district="NON_EXISTENT")
        assert len(cases) == 0


# ── ExportService ───────────────────────────────────────────────


class TestExportService:
    def _create_case(self, db) -> int:
        repo = CaseRepository(db)
        case = _make_case()
        return repo.create_case(case)

    def test_export_to_excel_creates_file(self, export_svc, db, tmp_path):
        self._create_case(db)
        output = tmp_path / "export.xlsx"
        result = export_svc.export_to_excel(output)
        assert result.exists()
        assert result.suffix == ".xlsx"

    def test_get_mis_data(self, export_svc, db):
        self._create_case(db)
        data = export_svc.get_mis_data()
        assert isinstance(data, dict)
        assert "total_cases" in data
        assert data["total_cases"] >= 1

    def test_get_mis_data_empty(self, export_svc):
        data = export_svc.get_mis_data()
        assert data["total_cases"] == 0

    def test_export_mis_to_excel(self, export_svc, db, tmp_path):
        self._create_case(db)
        output = tmp_path / "mis_export.xlsx"
        result = export_svc.export_mis_to_excel(output)
        assert result.exists()

    def test_export_to_excel_with_filter(self, export_svc, db, tmp_path):
        self._create_case(db)
        output = tmp_path / "filtered.xlsx"
        result = export_svc.export_to_excel(output, district="CVL")
        assert result.exists()

    def test_export_to_excel_with_progress(self, export_svc, db, tmp_path):
        self._create_case(db)
        progress = []
        output = tmp_path / "progress.xlsx"
        export_svc.export_to_excel(output, progress_cb=lambda c, t: progress.append((c, t)))
        assert output.exists()

    def test_export_to_excel_empty_db(self, export_svc, tmp_path):
        output = tmp_path / "empty.xlsx"
        result = export_svc.export_to_excel(output)
        assert result.exists()

    def test_export_mis_has_all_sheets(self, export_svc, db, tmp_path):
        import openpyxl
        self._create_case(db)
        output = tmp_path / "mis_full.xlsx"
        export_svc.export_mis_to_excel(output)
        wb = openpyxl.load_workbook(str(output))
        assert "Overview" in wb.sheetnames
        assert "By Status" in wb.sheetnames
        assert "By District" in wb.sheetnames
        assert "By Zone" in wb.sheetnames
        assert "By Work Type" in wb.sheetnames
        assert "Monthly Trend" in wb.sheetnames
        wb.close()

    def test_get_mis_data_keys(self, export_svc, db):
        self._create_case(db)
        data = export_svc.get_mis_data()
        expected_keys = [
            "district_status_counts", "total_cases", "total_amount",
            "total_amount_formatted", "by_status", "by_district",
            "by_zone", "by_work_type", "amount_by_district",
            "amount_by_status", "monthly_trend",
        ]
        for key in expected_keys:
            assert key in data

    def test_export_to_excel_status_filter(self, export_svc, db, tmp_path):
        repo = CaseRepository(db)
        c1 = repo.create_case(_make_case(order_no="60001001"))
        c2 = repo.create_case(_make_case(order_no="60001002"))
        repo.update_status(c2, CaseStatus.APPROVED)
        output = tmp_path / "approved.xlsx"
        export_svc.export_to_excel(output, status=CaseStatus.APPROVED)
        import openpyxl
        wb = openpyxl.load_workbook(str(output))
        ws = wb.active
        data_rows = ws.max_row - 1  # minus header
        wb.close()
        assert data_rows == 1


# ── TrackerService deep tests ───────────────────────────────────


class TestTrackerServiceDeep:
    @pytest.fixture
    def tracker_svc(self, db, tmp_path):
        from app.infrastructure.config_manager import ConfigManager
        ConfigManager._instance = None
        config = ConfigManager()
        config._config.set("General", "tracker_path", str(tmp_path / "tracker.xlsx"))
        svc = TrackerService(db=db, config=config)
        yield svc
        ConfigManager._instance = None

    def _create_case(self, db, **overrides) -> int:
        repo = CaseRepository(db)
        case = _make_case(**overrides)
        return repo.create_case(case)

    def test_resubmit_resets_to_pending(self, tracker_svc, db):
        case_id = self._create_case(db)
        tracker_svc.reject_case(case_id, "Wrong zone")
        tracker_svc.resubmit_case(case_id, "Fixed zone")
        case = tracker_svc.get_case(case_id)
        assert case.status == CaseStatus.PENDING

    def test_approve_syncs_to_tracker(self, tracker_svc, db, tmp_path):
        case_id = self._create_case(db)
        tracker_svc.approve_case(case_id)
        tracker_path = tmp_path / "tracker.xlsx"
        assert tracker_path.exists()

    def test_reject_stores_correction_details(self, tracker_svc, db):
        case_id = self._create_case(db)
        tracker_svc.reject_case(case_id, "Incorrect address")
        case = tracker_svc.get_case(case_id)
        assert case.correction_details == "Incorrect address"

    def test_update_tracker_full_sync(self, tracker_svc, db, tmp_path):
        self._create_case(db, order_no="60001001")
        self._create_case(db, order_no="60001002")
        tracker_svc.update_tracker()
        import openpyxl
        wb = openpyxl.load_workbook(str(tmp_path / "tracker.xlsx"))
        ws = wb["Sheet1"]
        data_rows = ws.max_row - 1  # minus header
        wb.close()
        assert data_rows == 2

    def test_get_case_by_order_no(self, tracker_svc, db):
        self._create_case(db, order_no="60099999")
        case = tracker_svc.get_case_by_order_no("60099999")
        assert case is not None
        assert case.order_no == "60099999"

    def test_list_cases_with_status_filter(self, tracker_svc, db):
        c1 = self._create_case(db, order_no="60001001")
        c2 = self._create_case(db, order_no="60001002")
        tracker_svc.approve_case(c2)
        # list_all expects status as a string (the enum .value)
        approved = tracker_svc.list_cases(status=CaseStatus.APPROVED.value)
        assert len(approved) >= 1
        assert all(c.status == CaseStatus.APPROVED for c in approved)


# ── ExportService deep tests ───────────────────────────────────


class TestExportServiceDeep:
    """Date filters, styling, and MIS edge cases."""

    @pytest.fixture(autouse=True)
    def setup(self, db, tmp_path):
        from app.infrastructure.config_manager import ConfigManager
        ConfigManager._instance = None
        config = ConfigManager()
        config._config.set("General", "output_dir", str(tmp_path / "output"))
        self.svc = ExportService(db=db, config=config)
        self.db = db
        self.tmp_path = tmp_path
        yield
        ConfigManager._instance = None

    def _create_case(self, **overrides) -> int:
        repo = CaseRepository(self.db)
        return repo.create_case(_make_case(**overrides))

    def test_export_with_date_from_filter(self):
        self._create_case(order_no="60001001")
        output = self.tmp_path / "date_from.xlsx"
        # date in the future should return no data rows
        result = self.svc.export_to_excel(output, date_from="2099-01-01")
        import openpyxl
        wb = openpyxl.load_workbook(str(result))
        ws = wb.active
        data_rows = ws.max_row - 1
        wb.close()
        assert data_rows == 0

    def test_export_with_date_to_filter(self):
        self._create_case(order_no="60001002")
        output = self.tmp_path / "date_to.xlsx"
        # date way in the past should return no data rows
        result = self.svc.export_to_excel(output, date_to="2000-01-01")
        import openpyxl
        wb = openpyxl.load_workbook(str(result))
        ws = wb.active
        data_rows = ws.max_row - 1
        wb.close()
        assert data_rows == 0

    def test_export_with_wide_date_range(self):
        self._create_case(order_no="60001003")
        output = self.tmp_path / "wide_range.xlsx"
        result = self.svc.export_to_excel(
            output, date_from="2020-01-01", date_to="2099-12-31"
        )
        import openpyxl
        wb = openpyxl.load_workbook(str(result))
        ws = wb.active
        data_rows = ws.max_row - 1
        wb.close()
        assert data_rows >= 1

    def test_export_header_styling(self):
        self._create_case(order_no="60001004")
        output = self.tmp_path / "styled.xlsx"
        self.svc.export_to_excel(output)
        import openpyxl
        wb = openpyxl.load_workbook(str(output))
        ws = wb.active
        # Header row should have blue fill
        header_cell = ws.cell(row=1, column=1)
        assert header_cell.font.bold is True
        assert header_cell.fill.start_color.rgb is not None
        wb.close()

    def test_export_with_zone_filter(self):
        self._create_case(order_no="60001005", zone="411")
        self._create_case(order_no="60001006", zone="507")
        output = self.tmp_path / "zone_filter.xlsx"
        result = self.svc.export_to_excel(output, zone="411")
        import openpyxl
        wb = openpyxl.load_workbook(str(result))
        ws = wb.active
        data_rows = ws.max_row - 1
        wb.close()
        assert data_rows >= 1

    def test_mis_empty_db(self):
        output = self.tmp_path / "mis_empty.xlsx"
        result = self.svc.export_mis_to_excel(output)
        assert result.exists()
        import openpyxl
        wb = openpyxl.load_workbook(str(result))
        assert "Overview" in wb.sheetnames
        wb.close()

    def test_mis_district_status_matrix(self):
        self._create_case(order_no="60001007", district="CVL")
        self._create_case(order_no="60001008", district="MDT")
        output = self.tmp_path / "mis_matrix.xlsx"
        self.svc.export_mis_to_excel(output)
        import openpyxl
        wb = openpyxl.load_workbook(str(output))
        assert "District × Status" in wb.sheetnames
        ws = wb["District × Status"]
        # Should have header + at least 2 district rows
        assert ws.max_row >= 3
        wb.close()

    def test_mis_data_aggregation(self):
        self._create_case(order_no="60001009", grand_total=100000)
        self._create_case(order_no="60001010", grand_total=200000)
        data = self.svc.get_mis_data()
        assert data["total_cases"] == 2
        assert data["total_amount"] >= 300000
